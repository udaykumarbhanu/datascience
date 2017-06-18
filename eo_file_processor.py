"""This pyhton code will read and process all the TAB, CSV, XLSX 
    and XLSM files present in below directory. 
"""
import os
import csv
import openpyxl
import xlrd
import db_connector
import time
import pandas as pd
import glob
import StringIO

__author__  = "Uday Bhanu"
__version__ = "1.0.0"
__all__ = ["table_truncation", "xlsm_file_processing", "tab_file_processing", "csv_file_processing", 
            "xlsx_file_processing", "table_reindexing",]


# Dev Source file location.
# src_file_path = r"C:\Users\udabhanu\Desktop\Subhash\input_data"

# Prod Input file location.
src_file_path = r"C:\EO_CPN\InData"


# This method is needed to make the insertion faster.
def chunker(seq, size):
    return (seq[pos:pos + size] for pos in xrange(0, len(seq), size))


class FileData2Table(object):
    def __init__(self):
        self.conn = db_connector.DatabaseConnection()

        print "************************************************************************"
        print "**************Please wait, I am processing your files.******************"
        print "************************************************************************"
    
    
    # Truncate all the tables.
    def table_truncation(self):        
        table_list = ["EO_MDS", "CPN_PID", "EO_MDS_TOP_CPN", "FDMT", "BUBESubBE", "EO_ACT_DEMAND"]

        try:            
            curr = self.conn.get_cursor()

            for table_name in table_list:
                truncate_query = "TRUNCATE TABLE EO_CPN..{}".format(table_name)
                print "We are truncating table :", table_name

                curr.execute(truncate_query)

        except Exception as ex:
            raise ex
        finally:
            self.conn.free(curr) 


    # XLSM file processing. ***This file needs to be processed before tab files***.
    def xlsm_file_processing(self):
        flag = False

        insert_query = """ INSERT INTO EO_MDS ("PF", "CPN", "CPN_PF", "Component Cost", "CME&O Reserve Provision (Release) Amount", 
            "MPa Owned Nettable Qty", "12M On-order Qty + VMI (CPa Owned) Qty", "12M Demand Qty", 
            "12M MPa Owned + (On-order + VMI (CPa Owned)) Excess Qty") VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?) """

        xlsm_file_list = [file_name for file_name in os.listdir(src_file_path) if file_name.endswith(".xlsm")]
        # print "List of all the XLSM files are :", xlsm_file_list

        if not xlsm_file_list:
            print "xlsm file is missing!"
            return False

        try:
            curr = self.conn.get_cursor()

            for xlsm_file_one in xlsm_file_list:
                print "XLSM File under execution is :", xlsm_file_one

                work_book = xlrd.open_workbook(os.path.join(src_file_path, xlsm_file_one), on_demand=True)
                sheet = work_book.sheet_by_name("Data")

                if sheet:
                    xlsm_file_data_list = list()
                    # In xlrd package,  row starts from 1 and column starts from 0. And first 2 rows are headers.
                    # data_list = sheet.row_values(2) # This is to print worksheet header.
                    for row_num in xrange(3, sheet.nrows, 1):
                        data_list = sheet.row_values(row_num)
                        # print data_list[1], data_list[5], data_list[6], data_list[14], data_list[23], data_list[24], data_list[25], data_list[26], data_list[28], data_list[30], data_list[40]
                        
                        col_1 = data_list[1]
                        col_5 = data_list[5]
                        col_6 = data_list[6]
                        col_14 = data_list[14]

                        col_23 = data_list[23]
                        col_24 = data_list[24]

                        # To handle some of the empty cells.
                        if isinstance(data_list[23], unicode):
                            col_23 = 0.0
                        if isinstance(data_list[24], unicode):
                            col_24 = 0.0

                        col_25 = col_23 - col_24

                        col_26 = data_list[26]
                        col_28 = data_list[28]
                        col_30 = data_list[30]
                        col_40 = data_list[40]
                        # print col_1, col_5, col_6, col_14, col_25, col_26, col_28, col_30, col_40
                        
                        # Storing each row data into xlsm_file_data_list which we will use later for batch insertion.
                        xlsm_file_data_list.append((col_1, col_5, col_6, col_14, col_25, col_26, col_28, col_30, col_40))
                    # print xlsm_file_data_list
                    
                    for batch in chunker(xlsm_file_data_list, 1000):
                        curr.executemany(insert_query, batch)
                        self.conn.commit()

                    flag = True
                else:
                    print "'Data' worksheet is not present in the current workbook!"
                    flag = False

        except Exception as ex:
            self.conn.rollback()
            raise ex
        finally:
            self.conn.commit()
            self.conn.free(curr)
            return flag


    # TAB file processing.
    def tab_file_processing(self):
        cpn_pid_list = list()

        # Top 2500 CPN query.
        select_query_1 = """ SELECT TOP 2500 [CPN], [CME&O Reserve Provision (Release) Amount]
                         FROM EO_MDS
                         WHERE [CME&O Reserve Provision (Release) Amount] >0 
                         ORDER BY [CME&O Reserve Provision (Release) Amount] DESC """

        # Bottom 2500 CPN query.
        select_query_2 = """ SELECT TOP 2500 [CPN], [CME&O Reserve Provision (Release) Amount]
                         FROM EO_MDS
                         WHERE [CME&O Reserve Provision (Release) Amount] <0 
                         ORDER BY [CME&O Reserve Provision (Release) Amount] ASC """

        insert_query = """ INSERT INTO CPN_PID ("CPN", "PID", "Quantity_Per") VALUES (?, ?, ?) """

        tab_files = glob.glob(os.path.join(src_file_path, "*.tab"))
        
        if not tab_files:
            print "tab files are missing!"
            return

        try:
            curr = self.conn.get_cursor()
            select_query_1_res = curr.execute(select_query_1).fetchall()
            select_query_2_res = curr.execute(select_query_2).fetchall()
            # print "select_query_1_res", select_query_1_res
            # print "select_query_2_res", select_query_2_res

            eo_mds_cpn_list = select_query_1_res + select_query_2_res
            eo_mds_labels = ["CPN", "CME&O Reserve Provision (Release) Amount"]
            eo_mds_cpn_df = pd.DataFrame.from_records(eo_mds_cpn_list, columns=eo_mds_labels)
            # print eo_mds_cpn_df

            # Loading tab file data to pandas dataframe.
            df_from_each_tab_file = (pd.read_csv(one_tab_file, sep="\t") for one_tab_file in tab_files)
            concatenated_df = pd.concat(df_from_each_tab_file, ignore_index=True)  
            # print "concatenated_df :", concatenated_df  

            resulted_df = pd.merge(eo_mds_cpn_df, concatenated_df, how="inner", on=["CPN"])  

            del resulted_df["CME&O Reserve Provision (Release) Amount"]
            # print resulted_df.columns
            cpn_pid_list = resulted_df.values.tolist()
            # print result_list

            for batch in chunker(cpn_pid_list, 1000):
                curr.executemany(insert_query, batch)
                self.conn.commit()

        except Exception as ex:
            raise ex
        finally:
            self.conn.commit()
            self.conn.free(curr)


    # XLSX file processing.
    def xlsx_file_processing(self):
        xlsx_file_list = [file_name.lower() for file_name in os.listdir(src_file_path) if file_name.endswith(".xlsx")]
        # print "List of all the XLSX files are :", xlsx_file_list

        if not xlsx_file_list:
            print "xlsx files are missing!"
            return

        for xlsx_file in xlsx_file_list:
            print "XLSX File under execution is :", xlsx_file

            if "business entity mapping" in xlsx_file:
                try:
                    curr = self.conn.get_cursor()               
                    insert_query = """ INSERT INTO BUBESubBE ("Product Family", "Business Unit", "Internal Business Entity", 
                           "Internal Sub Business Entity") VALUES (?, ?, ?, ?) """

                    work_book = openpyxl.load_workbook(os.path.join(src_file_path, xlsx_file))
                    sheet = work_book.get_sheet_by_name("Internal Mapping")

                    if sheet:                        
                        # As row in Excel starts from 1 and it goes to 'sheet.max_row'. But we DO NOT NEED HEADER so row STARTS from 1.
                        # But python range is upto sheet.max_row but NOT inlcuding sheet.max_row.
                        result_dict = dict()
                        for r in xrange(1, sheet.max_row+1, 1): 
                            row_data = list()           
                            # For each row we need data from column 1, 2, 5 and 6.
                                        
                            column_data_1 = sheet.cell(row=r, column=1)
                            # print column_data_1.value,
                            row_data.append(column_data_1.value)

                            column_data_2 = sheet.cell(row=r, column=2)
                            # print column_data_2.value,
                            row_data.append(column_data_2.value)
                            
                            column_data_5 = sheet.cell(row=r, column=5)
                            # print column_data_5.value,
                            row_data.append(column_data_5.value)

                            column_data_6 = sheet.cell(row=r, column=6)
                            # print column_data_6.value
                            row_data.append(column_data_6.value)

                            # print row_data
                            if all(row_data):
                                one_row_data = ":".join(row_data[:3])
                                if one_row_data not in result_dict:
                                    result_dict[one_row_data] = row_data[:]
                                else:
                                    pass
                            else:
                                break

                        # print result_dict

                        for key, value in result_dict.items():
                            # print key, value
                            curr.execute(insert_query, (value[0], value[1], value[2], value[3]))

                except Exception as ex:
                    self.conn.rollback()
                    raise ex
                finally:
                    self.conn.commit()
                    self.conn.free(curr)
            
            if "E_and_O_scrub_report.xlsx".lower() in xlsx_file:
                try:
                    curr = self.conn.get_cursor() 
                    insert_query = """ INSERT INTO EO_ACT_DEMAND ("PLID", "Prev month Fcst-Actual", "P1 12M demand",
                           "P0 12M demand") VALUES (?, ?, ?, ?) """

                    work_book = openpyxl.load_workbook(os.path.join(src_file_path, xlsx_file))
                    sheet = work_book.get_sheet_by_name("Summary")

                    if sheet:                  
                        # As row in Excel starts from 1 and it goes to 'sheet.max_row'. But we DO NOT NEED HEADER so row STARTS from 3.
                        # But python range is upto sheet.max_row but NOT inlcuding sheet.max_row.
                        for r in xrange(3, sheet.max_row+1, 1):           
                            # For each row we need data from column 2, 3, 4 and 5.
                                        
                            column_data_2 = sheet.cell(row=r, column=2)
                            # print column_data_2.value                       

                            column_data_3 = sheet.cell(row=r, column=3)
                            # print column_data_3.value
                            
                            column_data_4 = sheet.cell(row=r, column=4)
                            # print column_data_4.value

                            column_data_5 = sheet.cell(row=r, column=5)
                            # print column_data_5.value

                            # print column_data_2.value, column_data_3.value, column_data_4.value, column_data_5.value
                            
                            if any([column_data_2.value, column_data_3.value, column_data_4.value, column_data_5.value]):
                                curr.execute(insert_query, (column_data_2.value, column_data_3.value, column_data_4.value, column_data_5.value))
                            else:
                                break

                except Exception as ex:
                    self.conn.rollback()
                    raise ex
                finally:
                    self.conn.commit()
                    self.conn.free(curr)

    
    # CSV file processing.
    def csv_file_processing(self):
        insert_query = """ INSERT INTO FDMT ("PID", "PLID", "SPLID", "SUB_GROUP", "PF", "BU") VALUES (?, ?, ?, ?, ?, ?) """
        
        csv_file_list = [file_name for file_name in os.listdir(src_file_path) if file_name.endswith(".csv")]
        # print "List of all the CSV files are: ", csv_file_list

        if not csv_file_list:
            print "csv file is missing!"
            return
        
        try:
            curr = self.conn.get_cursor()

            for csv_file_one in csv_file_list:
                print "CSV File under execution is :", csv_file_one            

                with open(os.path.join(src_file_path, csv_file_one)) as csv_fobj:
                    csv_data = csv.reader(csv_fobj, delimiter = ",")
                    next(csv_data) # Header is not needed.

                    for one_line_csv_data in csv_data:
                        # print one_line_csv_data                        
                        
                        if any(one_line_csv_data):
                            # print one_line_csv_data[0], one_line_csv_data[2], one_line_csv_data[3], one_line_csv_data[4], one_line_csv_data[5], one_line_csv_data[6]
                            
                            col_0 = one_line_csv_data[0].decode('windows-1252') 
                            col_2 = one_line_csv_data[2].decode('windows-1252') 
                            col_3 = one_line_csv_data[3].decode('windows-1252')
                            col_4 = one_line_csv_data[4].decode('windows-1252')
                            col_5 = one_line_csv_data[5].decode('windows-1252')
                            col_6 = one_line_csv_data[6].decode('windows-1252')                          
                            
                            # curr.execute(insert_query, (one_line_csv_data[0], one_line_csv_data[2], one_line_csv_data[3], one_line_csv_data[4], one_line_csv_data[5], one_line_csv_data[6]))
                            curr.execute(insert_query, (col_0, col_2, col_3, col_4, col_5, col_6))                    
                        else:
                            break

                self.conn.commit()

        except Exception as ex:
            self.conn.rollback()
            raise ex
        finally:
            self.conn.commit()
            self.conn.free(curr)


    # Reindexing tables.
    def table_reindexing(self):
        cpn_pid_index_cpn_drop_query = "DROP INDEX [CPN_PID_INDEX_CPN] ON [EO_CPN].[dbo].[CPN_PID] WITH (ONLINE=OFF)"
        cpn_pid_index_pid_drop_query = "DROP INDEX [CPN_PID_INDEX_PID] ON [EO_CPN].[dbo].[CPN_PID] WITH (ONLINE=OFF)"
        cpn_pid_drop_query = "DROP STATISTICS [dbo].[CPN_PID].[CPN_PID_Stats]"
               
        cpn_pid_index_cpn_create_query = "CREATE INDEX [CPN_PID_INDEX_CPN] ON [EO_CPN].[dbo].[CPN_PID] ([CPN])"
        cpn_pid_index_pid_create_query = "CREATE INDEX [CPN_PID_INDEX_PID] ON [EO_CPN].[dbo].[CPN_PID] ([PID])"
        cpn_pid_create_query = "CREATE STATISTICS [CPN_PID_Stats] ON [EO_CPN].[dbo].[CPN_PID]([CPN], [PID])"

        try:            
            curr = self.conn.get_cursor()

            print "Now reindexing CPN_PID_INDEX_CPN!"
            curr.execute(cpn_pid_index_cpn_drop_query)
            curr.execute(cpn_pid_index_cpn_create_query)
            
            print "Now reindexing CPN_PID_INDEX_PID!"
            curr.execute(cpn_pid_index_pid_drop_query)
            curr.execute(cpn_pid_index_pid_create_query)

            print "Now reindexing STATISTICS!"
            curr.execute(cpn_pid_drop_query) 
            curr.execute(cpn_pid_create_query)           

        except Exception as ex:
            raise ex
        finally:
            self.conn.free(curr)
            if self.conn:
                self.conn.disconnect()


# All the methods are called here.
if __name__ == "__main__":
    filedata2table_obj = FileData2Table()

    filedata2table_obj.table_truncation()
    
    print "------------------------------------------------------------------------------"
    print "XLSM file processing STARTS at :", time.ctime()
    flag_result = filedata2table_obj.xlsm_file_processing()
    print "XLSM file processing ENDS at :", time.ctime()
    
    if flag_result:
        print "------------------------------------------------------------------------------"
        print "TAB files processing STARTS at :", time.ctime()
        filedata2table_obj.tab_file_processing()
        print "TAB files processing ENDS at :", time.ctime()
    else: 
        print "Dude to error in XLSM file processing, tab files are not processed!"

    print "------------------------------------------------------------------------------"
    print "CSV file processing STARTS at :", time.ctime()
    filedata2table_obj.csv_file_processing()
    print "CSV file processing ENDS at :", time.ctime()

    print "------------------------------------------------------------------------------"
    print "XLSX files processing STARTS at :", time.ctime()
    filedata2table_obj.xlsx_file_processing()
    print "XLSX files processing ENDS at :", time.ctime()

    print "------------------------------------------------------------------------------"
    print "Table reindexing and statistics recreation STARTS at :", time.ctime()
    filedata2table_obj.table_reindexing()
    print "Table reindexing and statistics recreation ENDS at :", time.ctime()

    print "***********************************************************************"
    print "All the file processed!!! Plz. go to database for processed data!"
    print "***********************************************************************"
   
    print "------------------------------------------------------------------------------"

  